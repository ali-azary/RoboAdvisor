import pandas as pd
import numpy as np
from scipy.optimize import minimize
import yfinance as yf
from openpyxl import load_workbook

# etfdata = pd.read_csv('screener-etf.csv')
etfdata = pd.read_excel('ETF-data.xlsx', sheet_name='ETF data',engine='openpyxl', index_col=[0])

# top volume ETFs
etfdata_sorted = etfdata.sort_values(by='Avg. Volume', ascending=False)
top = etfdata_sorted.head(100)
symvol=list(top.index)

# top Assets ETFs
etfdata_sorted = etfdata.sort_values(by='Assets', ascending=False)
top = etfdata_sorted.head(100)
symass=list(top.index)

# common ETFs between top volume and top assets
symbols=np.intersect1d(symvol,symass)


# download monthly price history for the past 5 years
prices=pd.DataFrame(columns=['date'])
start_date = "2018-01-01"
end_date = "2022-12-31"
symbol=symbols[0]
data = yf.download(symbol, start=start_date, end=end_date, interval="1mo")
prices['date'] = data.index
prices.set_index('date', inplace=True)
prices[symbol] = data['Adj Close']

for symbol in symbols[1:]:
    data = yf.download(symbol, start=start_date, end=end_date, interval="1mo")
    prices[symbol] = data['Adj Close']

prices = prices.dropna(axis=1)
 
# prices.to_csv('prices.csv')
book = load_workbook('ETF-data.xlsx')
writer = pd.ExcelWriter('ETF-data.xlsx', engine = 'openpyxl')
writer.book = book
prices.to_excel(writer, sheet_name = 'price history')
writer.close()

returns = prices.pct_change()


for symbol in returns.columns:
    returns[symbol]=returns[symbol]-etfdata['Exp. Ratio'][symbol]/12

    
book = load_workbook('ETF-data.xlsx')
writer = pd.ExcelWriter('ETF-data.xlsx', engine = 'openpyxl')
writer.book = book
returns.to_excel(writer, sheet_name = 'returns')
writer.close()